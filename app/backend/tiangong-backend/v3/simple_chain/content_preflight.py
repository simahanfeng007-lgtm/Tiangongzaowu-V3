"""Office 内容预检（P19-R2 M2 自 kernel.py 机械迁移，行为零变化）。

历史上位于 simple_chain/kernel.py 的 `_office_content_gaps`。本模块
只负责同轮纠错用的快速本地预检（读 host path、正则启发式、异常
静默跳过），不是 Gateway 权威验证器；权威内容判定见
src/total_gateway/outcome_oracles/artifact_content.py（读 immutable
object store bytes）。两者分权：本模块不得生成权威 predicate。
"""

from __future__ import annotations

import re
from pathlib import Path


def _office_content_gaps(
    user_message: str,
    generated_attachments: list[dict[str, str]] | None,
) -> list[str]:
    """内容级验收：office 交付物不得是空壳。

    真机 2026-08-29 复现：docx 只有标题、xlsx 只有一个空单元格，模型
    谎报内容详实（连人名分数都编好了），而存在性/完整性质检全绿。
    这里按用户请求推导内容下限（列名/行数/条数），对已生成的
    .xlsx/.docx 做无副作用读取核对；缺口进入纠错循环让模型补写，
    补不了则诚实报未完成。读取失败不产生缺口（损坏由 zip 质检负责）。
    """
    if not generated_attachments:
        return []
    message = str(user_message or "")
    gaps: list[str] = []
    rows_match = re.search(r"(\d+)\s*行", message)
    requested_rows = int(rows_match.group(1)) if rows_match else 0
    columns_match = re.search(
        r"([\u4e00-\u9fffA-Za-z0-9]+(?:\s*[、,，]\s*[\u4e00-\u9fffA-Za-z0-9]+)+)\s*列",
        message,
    )
    requested_columns: list[str] = []
    if columns_match:
        requested_columns = [
            re.sub(r"[两二三四五六七八九几0-9]+$", "", item.strip())
            for item in re.split(r"[、,，]", columns_match.group(1))
        ]
        requested_columns = [item for item in requested_columns if item]
    items_match = re.search(r"各?(\d+)\s*[条点项]", message)
    requested_items = int(items_match.group(1)) if items_match else 0
    for item in generated_attachments:
        if not isinstance(item, dict):
            continue
        path_text = str(item.get("path") or item.get("lujing") or "").strip()
        if not path_text:
            continue
        lowered = path_text.lower()
        name = Path(path_text).name
        try:
            if lowered.endswith(".xlsx"):
                import openpyxl

                workbook = openpyxl.load_workbook(path_text, read_only=True, data_only=True)
                sheet = workbook.active
                max_row = int(sheet.max_row or 0)
                max_col = int(sheet.max_column or 0)
                header: list[str] = []
                data_rows: list[tuple[str, ...]] = []
                if max_row >= 1:
                    rows_iter = sheet.iter_rows(values_only=True)
                    first = next(rows_iter, ())
                    header = [str(cell or "").strip() for cell in first]
                    for row in rows_iter:
                        cells = tuple(str(cell or "").strip() for cell in row)
                        if any(cells):
                            data_rows.append(cells)
                workbook.close()
                if max_row <= 1 and max_col <= 1:
                    gaps.append(f"内容缺失：{name} 是空表，需要写入表头与数据行")
                    continue
                if len(data_rows) >= 2 and len(set(data_rows)) == 1:
                    gaps.append(
                        f"内容缺失：{name} 的数据行全部相同（{('、'.join(data_rows[0]))}），"
                        "是占位符不是真实数据，需要写入各行不同的内容"
                    )
                    continue
                if requested_rows and max_row < requested_rows:
                    gaps.append(
                        f"内容缺失：{name} 需要约{requested_rows}行数据，实际只有{max_row}行"
                    )
                if requested_columns:
                    missing_columns = [
                        column
                        for column in requested_columns
                        if not any(column and column in cell for cell in header if cell)
                    ]
                    if missing_columns:
                        gaps.append(
                            f"内容缺失：{name} 缺少列 {('、'.join(missing_columns))}"
                        )
            elif lowered.endswith(".docx"):
                from docx import Document as _DocxDocument

                document = _DocxDocument(path_text)
                texts = [p.text.strip() for p in document.paragraphs if p.text.strip()]
                content_chars = sum(len(text) for text in texts)
                content_floor = max(30, requested_items * 16 if requested_items else 30)
                if content_chars < content_floor:
                    gaps.append(
                        f"内容缺失：{name} 正文约{content_chars}字，疑似只有标题的空壳，"
                        f"需要写入实际内容（约{content_floor}字以上）"
                    )
        except Exception:
            continue
    return gaps
